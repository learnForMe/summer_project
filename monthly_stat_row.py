import openpyxl
import datetime
from datetime import timedelta
from calendar import mdays
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils import coordinate_from_string
from openpyxl.styles import Font

def add_month():
	wb=load_workbook('testing.xlsx',read_only = False, data_only = True)
	worksheet= wb.get_sheet_names()
	wb.active
	monthly_STAT=wb.get_sheet_by_name('Monthly_STAT')
	italic24Font = Font(size=15, italic=True)
	monthly_STAT['a1'].font=italic24Font
	monthly_STAT['a1']= "Month"
	row=monthly_STAT.max_row
	mydate = datetime.date.today()
	this_month= mydate.strftime("%B %Y")
	if monthly_STAT.cell('a%d' % row).value != this_month:
		row+=1
		monthly_STAT['a%d' % row]= this_month

	#month=monthly_STAT.max_row
	wb.save('testing.xlsx')




