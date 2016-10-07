
import openpyxl
import datetime
from datetime import timedelta
from calendar import mdays
from openpyxl import load_workbook
from dateutil import relativedelta
from openpyxl.compat import range
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils import coordinate_from_string

def add_column():
	wb=load_workbook('testing.xlsx')
	sheet = wb.get_sheet_by_name('Sheet')
	wb.active
	curr_col = sheet.max_column
	col_count = sheet.max_column
	col_count += 1
	pre_col =sheet.max_column -1
	curr_col=get_column_letter(curr_col)
	pre_col =get_column_letter(pre_col)
	new_col= get_column_letter(col_count)
	#print (new_col)
	today = datetime.date.today()
	first = today.replace(day=1)
	lastMonth = first - datetime.timedelta(days=1)
	lastMonth= lastMonth.strftime("%B %Y")
	next_month = "{:%B %Y}".format(datetime.date.today() + relativedelta.relativedelta(months=1))
	#next_month= "{:%B %Y}".format(today + timedelta(mdays[today.month]))
	month="{:%B %Y}".format(datetime.date.today())
	#print (next_month)
	#sheet['%s1' % curr_col] = month
	#print (lastMonth)


	if sheet.cell('%s1' % curr_col).value != month:
		sheet['%s1' % new_col] = next_month
	
	wb.save('testing.xlsx')
