#!/usr/bin/env python3

"""Smartcard CardRequest(Modified by Gary Tsai).

__author__ = Gary Tsai and Freddy Velez
For Veterans Lounge of John Jay College

__author__ = "http://www.gemalto.com"

Copyright 2001-2012 gemalto
Author: Jean-Daniel Aussel, mailto:jean-daniel.aussel@gemalto.com

This file is part of pyscard.

pyscard is free software; you can redistribute it and/or modify
it under the terms of the GNU Lesser General Public License as published by
the Free Software Foundation; either version 2.1 of the License, or
(at your option) any later version.

pyscard is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU Lesser General Public License for more details.

You should have received a copy of the GNU Lesser General Public License
along with pyscard; if not, write to the Free Software
Foundation, Inc., 51 Franklin St, Fifth Floor, Boston, MA  02110-1301  USA
"""

from __future__ import print_function
from smartcard.CardType import ATRCardType
from smartcard.pcsc.PCSCCardRequest import PCSCCardRequest
from smartcard.util import toHexString , toBytes
import openpyxl
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils import coordinate_from_string
from openpyxl.styles import Font
from header import header
from search_student import search_Student
#from art import art_schedule
from formular import formular
from exec import exe
from incre_col import increase_col
import alert
import re
import os
import time, threading 
import _thread
import smtplib


class CardRequest(object):
    """A CardRequest is used for waitForCard() invocations and specifies what
    kind of smart card an application is waited for.
    """

    def __init__(self, newcardonly=True, readers=None, cardType=None,
        cardServiceClass=None, timeout=None):
        """Construct new CardRequest.

        newcardonly:        if True, request a new card
                            default is False, i.e. accepts cards already
                            inserted

        readers:            the list of readers to consider for
                            requesting a card default is to consider all
                            readers

        cardType:           the smartcard.CardType.CardType to wait for;
                            default is smartcard.CardType.AnyCardType,
                            i.e. the request will succeed with any card

        cardServiceClass:   the specific card service class to create
                            and bind to the card default is to create
                            and bind a smartcard.PassThruCardService

        timeout:            the time in seconds we are ready to wait for
                            connecting to the requested card.  default
                            is to wait one second to wait forever, set
                            timeout to None
        """
        self.pcsccardrequest = PCSCCardRequest(newcardonly, readers,
            cardType, cardServiceClass, timeout)

    def getReaders(self):
        """Returns the list or readers on which to wait for cards."""
        return self.pcsccardrequest.getReaders()

    def waitforcard(self):
        """Wait for card insertion and returns a card service."""
        return self.pcsccardrequest.waitforcard()

    def waitforcardevent(self):
        """Wait for card insertion or removal."""
        return self.pcsccardrequest.waitforcardevent()
os.system("clear")
wb=load_workbook('testing.xlsx', data_only = True)
wb.active
worksheet= wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet')
_thread.start_new_thread(exe,())
_thread.start_new_thread(increase_col,())


while __name__ == '__main__':
    """Small sample illustrating the use of CardRequest.py."""

    cardtype = "3B 8F 80 01 80 4F 0C A0 00 00 03 06 40 00 00 00 00 00 00 28"
    length=10
    print('\t''------Tap card to SIGN IN-------')

   # cr = CardRequest(timeout=10, cardType=cardtype)
    cr = CardRequest(timeout=None, newcardonly=True)
    cs = cr.waitforcard()
    cs.connection.connect()
    card = toHexString(cs.connection.getATR())
    SELECT = [0xFF, 0xCA, 0x00, 0x00, 0x00]

    response, sw1, sw2 = cs.connection.transmit( SELECT)
   # print ('response: ', response, ' status words: ', "%x %x" % (sw1, sw2))
    texting = toHexString(response).replace(' ','')
    word_len=len(texting)
    if card == cardtype and length == word_len :
        header()
        search_Student(texting)
        cs.connection.disconnect()
        formular()

    else:
        cs = cr.waitforcard()
        cs.connection.disconnect()
        os.system ('clear')
        print (alert.stop)
        os.system ('say Try Again')

 
   
