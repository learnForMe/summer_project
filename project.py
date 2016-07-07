
import openpyxl
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter, column_index_from_string
#from openpyxl.cell import get_column_letter

f=open ('lastday.txt', 'r')
text=f.read()
texting=text.split('/n')

wb=load_workbook('testing.xlsx',read_only = False, data_only = True)
sheet = wb.get_sheet_by_name('Sheet1')
ws=wb.active

#print texting
#count=0
#row=1
column=1
col=get_column_letter(column)


max_row = sheet.max_row
insert_name=max_row+1
print insert_name
ws['%s%d' % (col,insert_name)] =str(texting)
wb.save('testing.xlsx')

# testing with row incrementing by 1


'''
while row:
	row+=1
	col=get_column_letter(column)
	ws['%s%d' % (col,row)] =str(texting)
	wb.save('testing.xlsx')
'''
# testing with inserting data to row incremented by 1


f.closed
