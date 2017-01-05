import openpyxl
import datetime
from datetime import timedelta
from calendar import mdays
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils import coordinate_from_string
from openpyxl.styles import Font
from dateutil import relativedelta

def formular():
    wb=load_workbook('testing.xlsx',read_only = False, data_only = True)
    worksheet= wb.get_sheet_names()
    wb.active
    sheet = wb.get_sheet_by_name('Sheet')
    next_month = "{:%B %Y}".format(datetime.date.today() + relativedelta.relativedelta(months=1))
    curr_col = sheet.max_column
    curr_col=get_column_letter(curr_col)
    i=1
    if "Monthly_STAT" not in worksheet:
        wb.create_sheet (index=2, title="Monthly_STAT")
        monthly_STAT=wb.get_sheet_by_name('Monthly_STAT')
        monthly_STAT.column_dimensions['A'].width = 25
        italic24Font = Font(size=15, italic=True)
        monthly_STAT['a1'].font=italic24Font
        monthly_STAT['a1']= "Month"
        mydate = datetime.date.today()
        this_month= mydate.strftime("%B %Y")
        row=monthly_STAT.max_row 
        if monthly_STAT.cell('a%d' % row).value != this_month:
            row+=1
            monthly_STAT['a%d' % row]= this_month
        max_col=sheet.max_column
        max_row = sheet.max_row
        formular_col=get_column_letter(max_col)
        if monthly_STAT.cell('a%d' % row).value == sheet.cell('%s1' % curr_col).value:
            while row >=2 and max_col >=3:
                formular_col=get_column_letter(max_col)
                monthly_STAT['b%s' % row]= '=SUM(Sheet!%s2:%s%d)' % (formular_col,formular_col,max_row)
                row-=1
                max_col-=1
        '''         
        else:
            row = row -1
            max_col=sheet.max_column-1
            formular_col=get_column_letter(max_col)
            monthly_STAT['b%s' % row]= '=SUM(Sheet!%s2:%s%d)' % (formular_col,formular_col,max_row)
        '''
        wb.save('testing.xlsx')
        

    else:
        monthly_STAT=wb.get_sheet_by_name('Monthly_STAT')
        monthly_STAT.column_dimensions['A'].width = 25
        italic24Font = Font(size=15, italic=True)
        monthly_STAT['a1'].font=italic24Font
        monthly_STAT['a1']= "Month"
        mydate = datetime.date.today()
        this_month= mydate.strftime("%B %Y")
        row=monthly_STAT.max_row 
        if monthly_STAT.cell('a%d' % row).value != this_month:
            row+=1
            monthly_STAT['a%d' % row]= this_month 
        max_col=sheet.max_column
        max_row = sheet.max_row
        #formular_col=get_column_letter(max_col)

        if monthly_STAT.cell('a%d' % row).value == sheet.cell('%s1' % curr_col).value:
           while row >=2 and max_col >=3:
                formular_col=get_column_letter(max_col)
                monthly_STAT['b%s' % row]= '=SUM(Sheet!%s2:%s%d)' % (formular_col,formular_col,max_row)
                row-=1
                max_col-=1
        '''
            row = row -1
            max_col=sheet.max_column-1
            formular_col=get_column_letter(max_col)
            monthly_STAT['b%s' % row]= '=SUM(Sheet!%s2:%s%d)' % (formular_col,formular_col,max_row)
        else:
            row = row -1
            max_col=sheet.max_column-1
            formular_col=get_column_letter(max_col)
            monthly_STAT['b%s' % row]= '=SUM(Sheet!%s2:%s%d)' % (formular_col,formular_col,max_row)
            '''

        wb.save('testing.xlsx')
        





