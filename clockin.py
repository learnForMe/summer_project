import openpyxl
import datetime
from datetime import timedelta
from calendar import mdays
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils import coordinate_from_string
from openpyxl.styles import Font
from dateutil import relativedelta

class timesheet:
	def clockin(self):
		time=datetime.datetime.now().strftime("%Y/%m/%d %H:%M")
		
		return time

	def excel(self, a):
		wb = openpyxl.load_workbook("clocking.xlsx")
		sheet =wb.get_sheet_by_name('Sheet')
		sheet= wb.active
		sheet.column_dimensions['a'].width = 40
		sheet.column_dimensions['b'].width = 40
		sheet.column_dimensions['c'].width = 40
		sheet.column_dimensions['d'].width = 40
		sheet.column_dimensions['e'].width = 40
		sheet.column_dimensions['f'].width = 40
		sheet.column_dimensions['g'].width = 40
		sheet.column_dimensions['h'].width = 40
		#sheet["a1"]= str(self.register())
		next_row= sheet.max_column
		self.ids = a
		
		if self.ids == sheet.cell("a2").value:
			if "IN" in sheet.cell("a%d" % (self.column_to_add("a")-1)).value:
				sheet["a%d"% self.column_to_add("a")]= "OUT -> " + self.clockin()
			else:
				sheet["a%d"% self.column_to_add("a")]= "IN -> " + self.clockin()
		elif self.ids == sheet.cell("b2").value:
			if "IN" in sheet.cell("b%d" % (self.column_to_add("b")-1)).value:
				sheet["b%d"% self.column_to_add("b")]= "OUT -> " + self.clockin()
			else:
				sheet["b%d"% self.column_to_add("b")]= "IN -> " + self.clockin()
			#sheet["b%d"% self. column_to_add("b")]= self.clockin()
		
		elif self.ids == sheet.cell("c2").value:
			if "IN" in sheet.cell("c%d" % (self.column_to_add("c")-1)).value:
				sheet["c%d"% self.column_to_add("c")]= "OUT -> " + self.clockin()
			else:
				sheet["c%d"% self.column_to_add("c")]= "IN -> " + self.clockin()
			#sheet["c%d"% self. column_to_add("c")]= self.clockin()
		
		elif self.ids == sheet.cell("d2").value:
			if "IN" in sheet.cell("d%d" % (self.column_to_add("d")-1)).value:
				sheet["d%d"% self.column_to_add("d")]= "OUT -> " + self.clockin()
			else:
				sheet["d%d"% self.column_to_add("d")]= "IN -> " + self.clockin()
			#sheet["d%d"% self. column_to_add("d")]= self.clockin()	
		
		elif self.ids == sheet.cell("e2").value:
			if "IN" in sheet.cell("e%d" % (self.column_to_add("e")-1)).value:
				sheet["e%d"% self.column_to_add("e")]= "OUT -> " + self.clockin()
			else:
				sheet["e%d"% self.column_to_add("e")]= "IN -> " + self.clockin()
		'''		
		elif self.ids == sheet.cell("f2").value:
			if "IN" in sheet.cell("f%d" % (self.column_to_add("f")-1)).value:
				sheet["f%d"% self.column_to_add("f")]= "OUT -> " + self.clockin()
			else:
				sheet["f%d"% self.column_to_add("f")]= "IN -> " + self.clockin()
		elif self.ids == sheet.cell("g2").value:
			if "IN" in sheet.cell("g%d" % (self.column_to_add("g")-1)).value:
				sheet["g%d"% self.column_to_add("g")]= "OUT -> " + self.clockin()
			else:
				sheet["g%d"% self.column_to_add("g")]= "IN -> " + self.clockin()
		elif self.ids == sheet.cell("h2").value:
			if "IN" in sheet.cell("h%d" % (self.column_to_add("h")-1)).value:
				sheet["h%d"% self.column_to_add("h")]= "OUT -> " + self.clockin()
			else:
				sheet["h%d"% self.column_to_add("h")]= "IN -> " + self.clockin()
				'''		
		wb.save("clocking.xlsx")

	def register(self,x,y):
		wb = openpyxl.load_workbook("clocking.xlsx")
		sheet =wb.get_sheet_by_name('Sheet')
		sheet= wb.active
		
		self.name = x
		self.ids= y
		#while count <= 4:

		if sheet.cell('%s1'% get_column_letter(1)).value==None:
			sheet["%s1"% get_column_letter(1)]= self.name
			sheet["%s2"% get_column_letter(1)]= self.ids
		elif sheet.cell('%s1'% get_column_letter(2)).value==None and self.ids != sheet.cell("a2").value:
			sheet["%s1"% get_column_letter(2)]= self.name
			sheet["%s2"% get_column_letter(2)]= self.ids

		elif sheet.cell('%s1'% get_column_letter(3)).value==None and self.ids != sheet.cell("a2").value and self.ids != sheet.cell("b2").value:
			sheet["%s1"% get_column_letter(3)]= self.name
			sheet["%s2"% get_column_letter(3)]= self.ids
		
		elif sheet.cell('%s1'% get_column_letter(4)).value==None and self.ids != sheet.cell("a2").value and self.ids != sheet.cell("b2").value and self.ids != sheet.cell("c2").value and self.ids != sheet.cell("d2").value:
			sheet["%s1"% get_column_letter(4)]= self.name
			sheet["%s2"% get_column_letter(4)]= self.ids
		
		elif sheet.cell('%s1'% get_column_letter(5)).value==None and self.ids != sheet.cell("a2").value and self.ids != sheet.cell("b2").value and self.ids != sheet.cell("c2").value  and self.ids != sheet.cell("d2").value and self.ids != sheet.cell("e2").value:
			sheet["%s1"% get_column_letter(5)]= self.name
			sheet["%s2"% get_column_letter(5)]= self.ids
		'''	
		elif sheet.cell('%s1'% get_column_letter(6)).value==None and self.ids != sheet.cell("a2").value and self.ids != sheet.cell("b2").value and self.ids != sheet.cell("c2").value and self.ids != sheet.cell("d2").value and self.ids != sheet.cell("e2").value and self.ids != sheet.cell("f2").value:
			sheet["%s1"% get_column_letter(6)]= self.name
			sheet["%s2"% get_column_letter(6)]= self.ids
		elif sheet.cell('%s1'% get_column_letter(7)).value==None and self.ids != sheet.cell("a2").value and self.ids != sheet.cell("b2").value and self.ids != sheet.cell("c2").value and self.ids != sheet.cell("d2").value and self.ids != sheet.cell("e2").value and self.ids != sheet.cell("f2").value and self.ids != sheet.cell("g2").value: 
			sheet["%s1"% get_column_letter(7)]= self.name
			sheet["%s2"% get_column_letter(7)]= self.ids
		elif sheet.cell('%s1'% get_column_letter(8)).value==None and self.ids != sheet.cell("a2").value and self.ids != sheet.cell("b2").value and self.ids != sheet.cell("c2").value and self.ids != sheet.cell("d2").value and self.ids != sheet.cell("e2").value and self.ids != sheet.cell("f2").value and self.ids != sheet.cell("g2").value and self.ids != sheet.cell("h2").value:
			sheet["%s1"% get_column_letter(8)]= self.name
			sheet["%s2"% get_column_letter(8)]= self.ids
			'''
		wb.save("clocking.xlsx")
		
	
	def column_to_add(self,col): 
		wb = openpyxl.load_workbook("clocking.xlsx")
		sheet =wb.get_sheet_by_name('Sheet')
		sheet= wb.active
		sheet_max_row = sheet.max_row
		cell_coord = col + str(sheet_max_row)
		#print (sheet.cell(cell_coord).value)
		while sheet.cell(cell_coord).value == None:
			sheet_max_row -= 1
			cell_coord = col + str(sheet_max_row)
		sheet_max_row += 1
		return int(sheet_max_row)	

	#def grep(self,c,d):
'''
def main():
	
	first=timesheet()
	first.clockin()
	first.register()
	first.excel()
	

if __name__=='__main__':
	main()
'''
