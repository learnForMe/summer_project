import openpyxl
import datetime
from datetime import timedelta
from calendar import mdays
from openpyxl import load_workbook
from dateutil import relativedelta
from openpyxl.compat import range
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils import coordinate_from_string

def increase_col():
	wb=load_workbook('report.xlsx')
	sheet = wb.get_sheet_by_name('Sheet')
	wb.active
	curr_col = sheet.max_column
	col_count = sheet.max_column
	col_count += 1
	pre_col =sheet.max_column -1
	curr_col=get_column_letter(curr_col)
	#pre_col =get_column_letter(pre_col)
	new_col= get_column_letter(col_count)
	
	today = datetime.date.today()
	first = today.replace(day=1)
	lastMonth = first - datetime.timedelta(days=1)
	lastMonth= lastMonth.strftime("%B %Y")
	next_month = "{:%B %Y}".format(datetime.date.today() + relativedelta.relativedelta(months=1))
	
	month="{:%B %Y}".format(datetime.date.today())
	
	while sheet.cell('%s1' % curr_col).value == month:
		pass
	else:
		sheet['%s1' % new_col] = month
		
	
	wb.save('report.xlsx')

#increase_col()