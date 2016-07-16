import openpyxl

from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils import coordinate_from_string
import re

#insert_name=max_row+1
#print insert_name

def remove(y):
	#cell_location=re.sub(r'<','',y)
	#cell_location=re.sub(r'Cell','',cell_location)
	cell_location=re.sub(r'[< Cell Sheet1.]','',y)
	cell_location=re.sub(r'>','',cell_location)
	cell_location=re.sub(r'[A-Z]','',cell_location)

	#cell_location=re.sub(r'[< Cell Sheet1]','',cell_location)
	return cell_location


#row=1
#column=4
def search_Student (x):
	count =0
	first_time =1
	for row in sheet.iter_rows():
		for cell in row:
			data = cell.value
			
			if data == str(x):
				cell = str(cell)
				cool=int(remove(cell))
				stop_by= ws.cell('%s%d' % (col3,cool)).value
				print stop_by
				
				ws['%s%d' % (col3,cool)] = stop_by+1
				wb.save('testing.xlsx')
				count+=1
				
	if count< 1:
		print "This Student is NEW"
		student_name=raw_input("Enter Name -> ")
		ws['%s%d' % (col2,insert_name)] =str(student_name)
		ws['%s%d' % (col,insert_name)] =str(x)
		ws['%s%d' % (col3,insert_name)] =first_time
		wb.save('testing.xlsx')
	else:
		print "Welcome Back!"

		return data



#def occurance (y):
	
	

f=open ('lastday.txt', 'r')
text=f.read()
texting=text.split('\n')
#texting ="g"

wb=load_workbook('testing.xlsx',read_only = False, data_only = True)
sheet = wb.get_sheet_by_name('Sheet1')
ws=wb.active

col=get_column_letter(1)# convert column number to letter and use for first column (ID card data)
col2=get_column_letter(2)# use for second column (Student name)
col3=get_column_letter(3)#use for third column(occurance)

max_row = sheet.max_row
insert_name=max_row+1
    
search_Student(texting)        	
#wb.save('testing.xlsx')

f.closed







