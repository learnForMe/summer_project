import openpyxl
import datetime
from datetime import timedelta
from calendar import mdays
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils import coordinate_from_string
from openpyxl.styles import Font
from dateutil import relativedelta

def header():
	italic24Font = Font(size=18, italic=False)
	wb=load_workbook('testing.xlsx',read_only = False, data_only = True)
	sheet =wb.get_sheet_by_name('Sheet')
	wb.active
	month="{:%B %Y}".format(datetime.date.today())
	today = datetime.date.today()
	sheet['a1'].font=italic24Font
	sheet['b1'].font=italic24Font
	sheet['c1'].font=italic24Font
	sheet.column_dimensions['A'].width = 20
	sheet.column_dimensions['B'].width = 30
	sheet.column_dimensions['C'].width = 30
	sheet['a1']="UID"
	sheet['b1']="Name"
	sheet['c1']="Email"
	start_col =4
	start_month=get_column_letter(start_col)
	sheet['%s1' % start_month] = "August 2016"
	
	curr_col = sheet.max_column
	new_col= curr_col+1
	first = today.replace(day=1)
	new_col = get_column_letter(new_col)
	lastMonth = first - datetime.timedelta(days=1)
	lastMonth= lastMonth.strftime("%B %Y")
	#print (curr_col)
	this_month=get_column_letter(curr_col)
	if sheet.cell('%s1' % this_month).value == lastMonth:
		sheet['%s1' % new_col] = month
	
	sheet.column_dimensions['%s' % this_month].width = 20
	wb.save('testing.xlsx')



