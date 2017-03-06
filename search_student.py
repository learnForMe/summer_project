import openpyxl
import re
import os
from time import sleep
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils import coordinate_from_string
from timer import prompt_with_timeout
from reporting import logs
import random
from clockin import timesheet
timing = timesheet()

def remove(y):
    
    cell_location=re.sub(r'[\D]','',y)
    return cell_location

def search_Student (x):

    wb=load_workbook('testing.xlsx',read_only = False, data_only = True)
    worksheet= wb.get_sheet_names()
    wb.active
    sheet = wb.get_sheet_by_name('Sheet')
    max_row = sheet.max_row
    insert_name=max_row+1

    max_col=sheet.max_column
    col=get_column_letter(1)# convert column number to letter and use for first column (ID card data)
    col2=get_column_letter(2)# use for second column (Student name)
    col3=get_column_letter(3)# use for third column (Student's email)
    col4=get_column_letter(max_col)#use for third column(occurance)
    #print (col4)
    count =0
    first_time =1
    for row in sheet.iter_rows():
        for cell in row:
            data = cell.value
            if data == str(x):
                cell = str(cell)
                cool=int(remove(cell))
                stop_by= sheet.cell('%s%d' % (col4,cool)).value
                if stop_by == None:
                    stop_by = 0
                #print (stop_by)
                sheet['%s%d' % (col4,cool)] = stop_by+1
                #print (sheet.cell('%s%d' % (col4,cool)).value)
                wb.save('testing.xlsx')
                count+=1
                again=sheet.cell('%s%d' % (col2,cool)).value
                email=sheet.cell('%s%d' % (col3,cool)).value            
    if count< 1 :
        print ('\a\a\a\a\a\a')
        os.system ('clear')
        #os.system ('echo "New Student"')
        print ("\t\t****** NEW STUDENT ******\n")
        student_name , student_email =prompt_with_timeout()
        
        sheet['%s%d' % (col2,insert_name)] =str(student_name)
        sheet['%s%d' % (col,insert_name)] =str(x)
        sheet['%s%d' % (col4,insert_name)] =first_time
        sheet['%s%d' % (col3,insert_name)] =str(student_email)
        logs(x,student_name)
        #timing.register(student_name, x)
        timing.excel(x)
        wb.save('testing.xlsx')
    else:
        print ('\a')
        logs(x,again)
       # timing.register(again, x)
        timing.excel(x)
        if again == None or  email == None or again == "None" or  email == "None":
            print ('\a\a\a\a\a\a')
            #student_name=input("Enter Name -> ")
            #student_email=input("Enter Student's email -> ")
            student_name , student_email =prompt_with_timeout()
            sheet['%s%d' % (col2,cool)] = str(student_name)
            sheet['%s%d' % (col3,cool)] =str(student_email)
            again = sheet.cell('%s%d' % (col2,cool)).value
            email = sheet.cell('%s%d' % (col3,cool)).value
            
            #print (x)
            wb.save('testing.xlsx')
        
        print ("\n\tWelcome Back! ",again,"\n")

    
    '''if x == "CFD893A460":
        os.system ('say Welcome Back %s' % again)
    elif x == "CFD8AA9A20":
        os.system ('say Hello Mr. President')
    else:
        randNum=random.randrange(17)
        if (randNum) == 0:
            os.system ('say %s. Did you just roll out of bed? You look like shit' % again)  # use for prank (aka April Foo)
        elif (randNum)== 1:
            os.system ('say %s. Your head is so big, you don’t have dreams you have movies. ' % again)
        elif (randNum) == 2:
            os.system ('say %s. I see you are having a bad day. Here is a straw, suck it up!' % again)
        elif (randNum) == 3:
            os.system ('say %s. Please do not study too hard. You will fail anyway' % again)
        elif (randNum) == 4:
            os.system ('say %s. I thought you were smart, Until you open your mouth.' % again)
        elif (randNum) == 5:
            os.system ('say %s. now everyone will looks at you awkwardly' % again)
        elif (randNum) == 6:
            os.system ('say %s. Is ok to be a Trump supporter' % again)
        elif (randNum) == 7:
            os.system ('say %s. People like you are the reason we have middle fingers' % again)
        elif (randNum)== 8:
            os.system ('say %s. When your parent dropped you off at the school when you were little, they got a ticket for littering' % again)
        elif (randNum) == 9:
            os.system ('say %s. It’s not that I’m smarter than you, its just that you’re dumber than everyone else' % again)
        elif (randNum) == 10:
            os.system ('say %s. are you from tennessee, because you are the only ten i see' % again)
        elif (randNum) == 11:
            os.system ('say %s. You made me wish i were not a robot' % again)
        elif (randNum) == 12:
            os.system ('say %s. is your nickname wifi, because i am feeling the connection' % again)
        elif (randNum) == 13:
            os.system ('say %s. Are you a keyboard? because you are just my type' % again)
        elif (randNum) == 14:
            os.system ('say %s. there is nothing about me that is micro. or soft' % again)
        elif (randNum) == 15:
            os.system ('say %s. if i freeze, it is not a computer virus. I was just stunned by your look' % again)
        elif (randNum) == 16:
            os.system ('say %s. i want to squeeze my hardware into your software without having to worry about getting a virus' % again)
            '''



      


