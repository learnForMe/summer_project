import openpyxl
import datetime
from datetime import timedelta
from calendar import mdays
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils import coordinate_from_string
from openpyxl.styles import Font
from dateutil import relativedelta
from incre_col import increase_col

def column_to_add(col): 
	wb=load_workbook('report.xlsx',read_only = False, data_only = True)
	sheet =wb.get_sheet_by_name('Sheet')
	sheet_max_row = sheet.max_row
	cell_coord = col + str(sheet_max_row)
	#print (sheet.cell(cell_coord).value)
	while sheet.cell(cell_coord).value == None:
		sheet_max_row -= 1
		cell_coord = col + str(sheet_max_row)
	sheet_max_row += 1
	return int(sheet_max_row)

def logs (a,b):
	wb=load_workbook('report.xlsx',read_only = False, data_only = True)
	sheet =wb.get_sheet_by_name('Sheet')
	wb.active
	mydate = datetime.date.today()
	this_month= mydate.strftime("%B %Y")
	curr_col= sheet.max_column
	curr_row = sheet.max_row
	min_col=sheet.min_column
	min_col=get_column_letter(min_col)
	curr_col = get_column_letter(curr_col)
	sheet['%s1' % curr_col] = this_month
	sheet.column_dimensions['%s' %  curr_col].width = 50
	time=datetime.datetime.now().strftime("%d %H:%M")
	if b != None:
		new_row=curr_row+1
		#sheet['%s%d' % (curr_col,new_row)] = a+ " " + b +" "+ time
		sheet['%s%d' % (curr_col,column_to_add(curr_col))] = time+">>> "+a+"   "+b
		
	else:
		new_row=curr_row+1
		b = " "
		sheet['%s%d' % (curr_col,column_to_add(curr_col))] = time+">>> "+a+"   "+b
		
	#print (column_to_add(curr_col))
	wb.save('report.xlsx')
	#sprint (time)

